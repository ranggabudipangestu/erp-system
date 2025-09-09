from typing import List, Dict, Any, Optional
from datetime import datetime, date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Excel export service for ERP entities"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
        
    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        sheet_name: str = "Data",
        headers: Optional[List[str]] = None
    ) -> BytesIO:
        """
        Export data to Excel format
        
        Args:
            data: List of dictionaries containing row data
            sheet_name: Name of the Excel sheet
            headers: Optional custom headers. If None, will use dict keys from first row
            
        Returns:
            BytesIO: Excel file content
        """
        if not data:
            raise ValueError("No data provided for export")
            
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = sheet_name
        
        # Determine headers
        if headers is None:
            headers = list(data[0].keys())
            
        self._write_headers(headers)
        self._write_data(data, headers)
        self._apply_formatting(headers)
        
        # Save to BytesIO
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        return output
    
    def _write_headers(self, headers: List[str]):
        """Write header row with formatting"""
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            cell.value = header.replace('_', ' ').title()
            
            # Header formatting
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _write_data(self, data: List[Dict[str, Any]], headers: List[str]):
        """Write data rows"""
        for row_num, row_data in enumerate(data, 2):  # Start from row 2
            for col_num, header in enumerate(headers, 1):
                cell = self.worksheet.cell(row=row_num, column=col_num)
                value = row_data.get(header, "")
                
                # Format different data types
                if isinstance(value, (datetime, date)):
                    cell.value = value.strftime("%Y-%m-%d %H:%M:%S") if isinstance(value, datetime) else value.strftime("%Y-%m-%d")
                elif isinstance(value, (int, float)):
                    cell.value = value
                else:
                    cell.value = str(value) if value is not None else ""
    
    def _apply_formatting(self, headers: List[str]):
        """Apply formatting to the worksheet"""
        # Auto-adjust column widths
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            
            # Calculate max width for column
            max_length = len(header.replace('_', ' ').title())
            for row in self.worksheet.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            
            # Set column width (with some padding)
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders and alternating row colors
        from openpyxl.styles import Border, Side
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders to all cells with data
        max_row = self.worksheet.max_row
        max_col = len(headers)
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = self.worksheet.cell(row=row, column=col)
                cell.border = thin_border
                
                # Alternating row colors (except header)
                if row > 1 and row % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


class InvoiceExcelExporter(ExcelExporter):
    """Specialized Excel exporter for invoice data"""
    
    def export_invoices(self, invoices: List[Dict[str, Any]]) -> BytesIO:
        """Export invoice data with custom formatting"""
        headers = [
            'invoice_id', 'invoice_number', 'customer_name', 'invoice_date',
            'due_date', 'subtotal', 'tax_amount', 'total_amount', 'status',
            'created_at', 'updated_at'
        ]
        
        return self.export_to_excel(
            data=invoices,
            sheet_name="Invoices",
            headers=headers
        )


class OrderExcelExporter(ExcelExporter):
    """Specialized Excel exporter for order data"""
    
    def export_orders(self, orders: List[Dict[str, Any]]) -> BytesIO:
        """Export order data with custom formatting"""
        headers = [
            'order_id', 'order_number', 'customer_name', 'order_date',
            'delivery_date', 'total_amount', 'status', 'payment_status',
            'created_at', 'updated_at'
        ]
        
        return self.export_to_excel(
            data=orders,
            sheet_name="Orders",
            headers=headers
        )


class InventoryExcelExporter(ExcelExporter):
    """Specialized Excel exporter for inventory/stock data"""
    
    def export_inventory(self, inventory: List[Dict[str, Any]]) -> BytesIO:
        """Export inventory data with custom formatting"""
        headers = [
            'product_id', 'product_name', 'sku', 'category',
            'current_stock', 'min_stock_level', 'unit_price',
            'total_value', 'last_updated', 'location'
        ]
        
        return self.export_to_excel(
            data=inventory,
            sheet_name="Inventory",
            headers=headers
        )